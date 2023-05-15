import pandas as pd
import datetime as dt
from datetime import datetime as dt2
import numpy as np
import yfinance as yf
from pandas_datareader import data as pdr
import scipy.optimize as sc
import plotly.graph_objects as go
import openpyxl as xl

yf.pdr_override()

class Color:
    """
    Класс добавляет названия цветов для оформления вывода текста
    """
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

def get_data(stocks, start, end):
    """
    Импортирует данные
    :param stocks: список акций
    :param start: время начальное
    :param end: время конечное
    :return: среднюю ежедневную доходность и ковариационную матрицу
    """
    stockdata = pdr.get_data_yahoo(stocks, start=start, end=end)
    stockdata = stockdata['Close']  # Получили цены закрытия
    files = ['United States 10-Year Bond Yield Historical Data.csv', 'Proshlye_dannye_po_SPY.csv']  # Список с нашими файлами
    returns = stockdata.pct_change()  # Получили дневную доходность


    # # ETF
    # df = pd.read_csv(files[1])  # Прочитали файл с данными по облигациям
    # df['Дата'] = df['Дата'].apply(lambda x: x.replace('.', '/'))  # заменили точки на палочки
    # df['Date'] = pd.to_datetime(df['Дата'], dayfirst=True)
    # df = df.set_index(df['Date'])
    # df = df['Изм. %'].apply(lambda x: x.replace(',', '.'))  # Заменили запятые на точки
    # df = df.apply(lambda x: float(x[0:-1]) if "%" in x else float(x))  # убрали знак процента
    # returns['ETF'] = df
    #
    # # Облигации
    # df = pd.read_csv(files[0])  # Прочитали файл с данными по облигациям)  # заменили точки на палочки
    # df['Date'] = pd.to_datetime(df['Date'])
    # df = df.set_index(df['Date'])
    # df = df['Change %'].apply(lambda x: float(x[0:-1]) if "%" in x else float(x))  # убрали знак процента
    # returns['Bound'] = df


    print(returns)
    mean = returns.mean()  # Получили среднее доходности
    covMatrix = returns.cov()  # Получили матрицу ковариации
    print(mean, covMatrix)
    return mean, covMatrix


def portfolioPerformance(weights, meanReturns, covMatrix):
    """
    Функция находит доходность портфеля и его среднеквадратичное отклонение.
    :param weights: веса активов.
    :param meanReturns: средняя доходность в сутки.
    :param covMatrix: матрица ковариации.
    :return: доходность за год, стандартное отклонение.
    """
    returns = np.sum(meanReturns*weights)*252
    std = np.sqrt(
            np.dot(weights.T, np.dot(covMatrix, weights))
           )*np.sqrt(252)
    return returns, std


def negativePP(weights, meanReturns, covMatrix):
    """
    Функция находит доходность портфеля.
    :param weights: веса активов.
    :param meanReturns: средняя доходность в сутки.
    :return: доходность за год.
    """
    returns = portfolioPerformance(weights, meanReturns, covMatrix)[0]
    return - returns


def maxPPerformance(meanReturns, covMatrix, riskFreeRate = 0, constraintSet=(0, 1)):
    """
    Минимизирует негативную доходность портфеля.
    :param meanReturns: средняя доходность активов.
    :param riskFreeRate: безрисковая ставка доходности.
    :param constraintSet: параметр необходимый для решения уравнения.
    :return: большой словарь по оптимизации заданных параметров.
    """
    numAssets = len(meanReturns) # количество активов
    args = (meanReturns, covMatrix)
    constraints = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
    bound = constraintSet
    bounds = tuple(bound for asset in range(numAssets))
    result = sc.minimize(negativePP, numAssets*[1./numAssets], args=args,
                         method='SLSQP', bounds=bounds, constraints=constraints)
    return result


def negativeSR(weights, meanReturns, covMatrix, riskFreeRate = 1):
    """
    В библиотеки scipy нет функции максимизации, зато есть функция минимизации. Так как мы хотим найти максимальный
    коэффициент Шарпа, то сначала мы сделаем его отрицательным, чтобы минимизировать отрицательный. Так мы найдем и
    положительный.
    :param weights: веса активов.
    :param meanReturns: средняя доходность активов.
    :param covMatrix: матрица ковариации.
    :param riskFreeRate: безрисковая ставка доходности.
    :return: негативный коэф шарпа.
    """
    pReturns, pStd = portfolioPerformance(weights, meanReturns, covMatrix)
    return - (pReturns - riskFreeRate)/pStd


def maxSRatio(meanReturns, covMatrix, riskFreeRate = 1, constraintSet=(0.04, 0.5)):
    """
    Минимизирует негативный К шарпа балансируя веса портфеля.
    :param meanReturns: средняя доходность активов.
    :param covMatrix: матрица ковариации.
    :param riskFreeRate: безрисковая ставка доходности.
    :param constraintSet: параметр необходимый для решения уравнения.
    :return: большой словарь по оптимизации заданных параметров .
    """
    numAssets = len(meanReturns)  # количество активов
    args = (meanReturns, covMatrix, riskFreeRate)
    constraints = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
    bound = constraintSet
    bounds = tuple(bound for asset in range(numAssets))
    result = sc.minimize(negativeSR, numAssets*[1./numAssets], args=args,
                        method='SLSQP', bounds=bounds, constraints=constraints)
    return result


def portfolioVariance(weights, meanReturns, covMatrix):
    """
    Выдает отклонение портфеля (вариацию)
    :param weights: веса
    :param meanReturns: среднее
    :param covMatrix: ковариация
    :return: отклонения
    """
    return portfolioPerformance(weights, meanReturns, covMatrix)[1]


def minimizeVariance(meanReturns, covMatrix, constraintSet=(0.04, 0.5)):
    """
    Выдает веса для минимальной вариации
    :param meanReturns: средняя доходность
    :param covMatrix: матрица ковариации
    :param constraintSet: параметрр для решения уравнения
    :return:
    """
    numAssets = len(meanReturns)
    args = (meanReturns, covMatrix)
    constraints = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
    bound = constraintSet
    bounds = tuple(bound for asset in range(numAssets))
    result = sc.minimize(portfolioVariance, numAssets*[1./numAssets], args=args,
                        method='SLSQP', bounds=bounds, constraints=constraints)
    return result

# def negativeMSR(weights, meanReturns, covMatrix, riskFreeRate = 4):


def calculatedResults(meanReturns, covMatrix, riskFreeRate=1, constraintSet=(0.04, 0.20)):
    """
    Общая функция, которая вызывает остальные функции расчета, обрабатывает их результат
    :param meanReturns: средняя
    :param covMatrix: матрица ковариации
    :param riskFreeRate: безрисковая ставка доходноости
    :param constraintSet: параметр для решения уравнения.
    :return:
    1. Доходность максимального к Шапра
    2. Волатильность максимального к Шарпа
    3. Веса портфеля максимальног к Шарпа
    4. Доходность минимальной волатильности
    5. Волотильность минимальной волатильности
    6. Веса минимальной волатильности
    7. Доходность промежуточных параметров для построения графика
    8. Валотильноость промежуточных параметров для построения графика
    """
    print(Color.BOLD + Color.BLUE + Color.UNDERLINE + 'Итог' + Color.END)
    # Максимальный к Шарпа
    maxSR_Portfolio = maxSRatio(meanReturns, covMatrix, riskFreeRate, constraintSet)
    maxSR_returns, maxSR_std = portfolioPerformance(maxSR_Portfolio['x'], meanReturns, covMatrix)
    maxSR_allocation = pd.DataFrame(maxSR_Portfolio['x'], index=meanReturns.index, columns=['allocation'])
    maxSR_allocation.allocation = [round(i * 100, 3) for i in maxSR_allocation.allocation]
    print(Color.GREEN + '\nМаксимальный к Шарпа, веса активов' + Color.END)
    print(maxSR_allocation)
    print(Color.DARKCYAN + '\nдоходность -' + Color.END, maxSR_returns,
          Color.DARKCYAN +  'волатильность - ' + Color.END,  maxSR_std)

    # Минимальная волатильность
    minVol_Portfolio = minimizeVariance(meanReturns, covMatrix, constraintSet)
    minVol_returns, minVol_std = portfolioPerformance(minVol_Portfolio['x'], meanReturns, covMatrix)
    minVol_allocation = pd.DataFrame(minVol_Portfolio['x'], index=meanReturns.index, columns=['allocation'])
    minVol_allocation.allocation = [round(i * 100, 3) for i in minVol_allocation.allocation]
    print(Color.GREEN + '\nМинимальная волатильность, веса активов' + Color.END)
    print(minVol_allocation)
    print(Color.DARKCYAN + '\nдоходность -' + Color.END, maxSR_returns,
          Color.DARKCYAN + 'волатильность - ' + Color.END, maxSR_std)

    # Максимальная доходность
    maxPerf_Portfolio = maxPPerformance(meanReturns, covMatrix, riskFreeRate, constraintSet)
    maxPerf_returns, maxPerf_std = portfolioPerformance(maxPerf_Portfolio['x'], meanReturns, covMatrix)
    maxPerf_allocation = pd.DataFrame(maxPerf_Portfolio['x'], index=meanReturns.index, columns=['allocation'])
    maxPerf_allocation.allocation = [round(i * 100, 0) for i in maxPerf_allocation.allocation]
    print(Color.GREEN + '\nМаксимальная доходность, веса активов' + Color.END)
    print(maxPerf_allocation)
    print(Color.DARKCYAN + '\nдоходность -' + Color.END, maxPerf_returns,
          Color.DARKCYAN + 'волатильность - ' + Color.END, maxPerf_std)

    # Граница эффективности
    efficientList = []
    targetReturns = np.linspace(minVol_returns, maxSR_returns, 15)
    for target in targetReturns:
        efficientList.append(efficientOpt(meanReturns, covMatrix, target)['fun'])
    return maxSR_returns, maxSR_std, maxSR_allocation, minVol_returns, minVol_std, \
        minVol_allocation, maxPerf_returns, maxPerf_std, maxPerf_allocation, efficientList, targetReturns


def portfolioReturn(weights, meanReturns, covMatrix):
    """Считает доходность портфеля"""
    return portfolioPerformance(weights, meanReturns, covMatrix)[0]


def efficientOpt(meanReturns, covMatrix, returnTarget, constraintSet=(0, 1)):
    """Считает параметры для построения границы эффективности"""
    numAssets = len(meanReturns)
    args = (meanReturns, covMatrix)
    constraints = ({'type': 'eq', 'fun': lambda x: portfolioReturn(x, meanReturns, covMatrix) - returnTarget},
                   {'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
    bound = constraintSet
    bounds = tuple(bound for asset in range(numAssets))
    effOpt = sc.minimize(portfolioVariance, numAssets * [1. / numAssets], args=args, method='SLSQP', bounds=bounds,
                         constraints=constraints)
    return effOpt


def EF_graph(meanReturns, covMatrix, riskFreeRate=1, constraintSet=(0.04, 0.3)):
    """Строит границу эффективности"""
    maxSR_returns, maxSR_std, maxSR_allocation, minVol_returns, minVol_std, minVol_allocation, \
        maxPerf_returns, maxPerf_std, maxPerf_allocation, efficientList, targetReturns = \
        calculatedResults(meanReturns, covMatrix, riskFreeRate, constraintSet)
    # Максимальный к Шарпа
    MaxSharpeRatio = go.Scatter(
        name='Максимальный к Шарпа',
        mode='markers',
        x=[round(maxSR_std, 4)],
        y=[round(maxSR_returns, 4)],
        marker=dict(color='red', size=14, line=dict(width=3, color='black'))
    )
    # Минимальная волатильность
    MinVol = go.Scatter(
        name='Минимальная волатильность',
        mode='markers',
        x=[round(minVol_std, 4)],
        y=[round(minVol_returns, 4)],
        marker=dict(color='green', size=14, line=dict(width=3, color='black'))
    )
    # Максимальная доходность
    MaxPP = go.Scatter(
        name='Максимальная доходность',
        mode='markers',
        x=[round(maxPerf_std, 4)],
        y=[round(maxPerf_returns, 4)],
        marker=dict(color='blue', size=14, line=dict(width=3, color='black'))
    )
    # Граница эффективности
    EF_curve = go.Scatter(
        name='Граница эффективности',
        mode='lines',
        x=[round(ef_std, 4) for ef_std in efficientList],
        y=[round(target, 4) for target in targetReturns],
        line=dict(color='black', width=4, dash='dashdot')
    )
    data = [MaxSharpeRatio, MinVol, EF_curve, MaxPP]
    layout = go.Layout(
        title='Оптимизация портфеля с границей эффективности',
        yaxis=dict(title='Годовой доход (%)'),
        xaxis=dict(title='Годовая волатильность (%)'),
        showlegend=True,
        legend=dict(
            x=0.75, y=0, traceorder='normal',
            bgcolor='#E2E2E2',
            bordercolor='black',
            borderwidth=2),
        width=800,
        height=600)

    fig = go.Figure(data=data, layout=layout)
    return fig.show()

# def writer():
#     """Функция записывает что-то в ексель файл, просто задел на будущее"""
#     book = xl.Workbook()
#     book.remove(book.active)
#
#     book.create_sheet('1')
#     book.create_sheet('2')
#     book.create_sheet('3')
#     a, b, c, d, e, f, g, h = calculatedResults(meant, cov, 4)
#     print(f)
#     for i in g:
#         i = float(i)
#     for i in h:
#         i = float(i)
#     l = [float(a), float(b), str(c), float(d), float(e), str(f)]
#     book.worksheets[0].append(l)
#     book.worksheets[1].append(list(g))
#     book.worksheets[2].append(list(h))
#
#     book.save('sample.xlsx')
#
stocklist = ['KO', 'CVX', 'MCD', 'V', 'HPQ', 'MCO', 'DVA', 'KR', 'MCK', 'AON']  # Лист акций

endDate = dt2(2020, 12, 31)
startDate = endDate - dt.timedelta(days=365)

weight = 10 * [1./10]
meant, cov = get_data(stocks=stocklist, start=startDate, end=endDate)


# print(get_data(stocklist, startDate, endDate))

# print(calculatedResults(meant, cov, 4))

#  Облигации
df = pd.read_csv('Proshlye_dannye_-_S_amp_amp_P_500.csv')  # Прочитали файл с данными по облигациям
df['Дата'] = df['Дата'].apply(lambda x: x.replace('.', '/'))  # заменили точки на палочки
df['Date'] = pd.to_datetime(df['Дата'], dayfirst=True)
df = df.set_index(df['Date'])
df = df['Изм. %'].apply(lambda x: x.replace(',', '.'))  # Заменили запятые на точки
market = df.apply(lambda x: float(x[0:-1]) if "%" in x else float(x))  # убрали знак процента
#
# print(fun(meant))
# print(meant)
# a, b, c, d, e, f, g, h = calculatedResults(meant, cov, 4)
# print(c)
# print(f)
calculatedResults(meant, cov, 1, (0.04, 0.25))
# EF_graph(meant, cov, 4, (0.04, 0.20))
